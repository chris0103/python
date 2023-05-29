import json
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

font_eng = Font(name='Verdana', size=10)
font_chn = Font(name='Microsoft YaHei', size=10)
font_site = Font(name='Verdana', size=10, bold=True)
fill_tag = PatternFill(fill_type='solid', fgColor='E2EFDA')
fill_site = PatternFill(fill_type='solid', fgColor='00FFCC99')

wb = Workbook()
ws = wb.active
ws.title = "HOP APIs"
ws.cell(row=1, column=1, value='Method').font = font_site
ws.cell(row=1, column=2, value='Endpoint').font = font_site
ws.cell(row=1, column=3, value='Summary').font = font_site
row = 2


def generate_api_table():
    global row
    base_path = swagger_json.get('basePath', '')
    last_tag = ''
    for path, path_info in swagger_json['paths'].items():
        endpoint = base_path + path
        for method, method_info in path_info.items():
            method = method.upper()
            summary = method_info.get('summary', '')
            tag = method_info['tags'][0]
            if last_tag != tag:
                print(tag)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                cell_tag = ws.cell(row=row, column=1, value=tag)
                cell_tag.font = font_eng
                cell_tag.fill = fill_tag
                row += 1
                last_tag = tag
            print(f'[{method}] {endpoint}\t{summary}')
            ws.cell(row=row, column=1, value=method).font = font_eng
            ws.cell(row=row, column=2, value=endpoint).font = font_eng
            ws.cell(row=row, column=3, value=summary).font = font_chn
            row += 1


api_sites = {'op-identity': 'http://10.31.141.141:30005/swagger/v1/swagger.json',
             'op-report': 'http://10.31.141.141:30006/swagger/v1/swagger.json',
             'op-wxapi': 'http://10.31.141.141:30001/swagger/v1/swagger.json',
             'op-wxcentral': 'http://10.31.141.141:30008/swagger/v1/swagger.json',
             'op-job': 'http://10.31.141.141:30010/swagger/v1/swagger.json'}

for key, value in api_sites.items():
    print(f'{key}: {value}')
    response = requests.get(value)
    swagger_json = json.loads(response.text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    site = ws.cell(row=row, column=1, value=key)
    site.font = font_site
    site.fill = fill_site
    row += 1
    generate_api_table()


# load local file when Swagger endpoint is blocked
# with open('/Users/i0480092/systems/WeChat OP/APIs/swagger/UAT-wxapi.json', 'r') as file:
#     swagger_json = json.load(file)
# generate_api_table()
#
# with open('/Users/i0480092/systems/WeChat OP/APIs/swagger/UAT-wxcentral.json', 'r') as file:
#     swagger_json = json.load(file)
# generate_api_table()
#
# ...

wb.save("HOP APIs.xlsx")
